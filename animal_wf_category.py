from openpyxl import load_workbook
from pymongo import MongoClient

client = MongoClient()
db = client["water_footprint"]
collection = db["animal_products_category"]
collection.delete_many({})

# Load excel file and select sheet
wb = load_workbook('./data/animal-products.xlsx')
ws = wb["animal"]

# Select all column and sub column labels
column_labels = ws[3]
sub_column_labels = ws[4]

# Rows we wish to loop over
rows = ws[5:257+2]


def create_product_list(excel_rows):
    """
    Creates a list of dictionaries representing an animal product and its water footprint.
    Water footprint is represented at a global average + country averages.
    Includes both green, blue, and grey water footprint types.
    :param excel_rows: Excel rows to iterate through
    :return: A list of animal products
    """
    products = []
    current_category = ""
    for idx, row in enumerate(excel_rows):
        product = {"countries": []}
        if row[0].value:
            for i in range(len(column_labels)):
                if column_labels[i].value == "Product discription (HS)":
                    product["product"] = row[i].value
                elif column_labels[i].value == "Product description (SITC)":
                    if row[i].value:
                        product["product_category"] = row[i].value
                        current_category = row[i].value
                    else:
                        product["product_category"] = current_category
                elif sub_column_labels[i].value == "Weighted average":
                    if column_labels[i-3].value == "World Average":
                        product["water_footprint_global_average"] = {
                            "green": row[i].value,
                            "blue": excel_rows[idx + 1][i].value,
                            "grey": excel_rows[idx + 2][i].value
                        }
                    else:
                        product["countries"].append(
                            {
                                "country": column_labels[i-3].value,
                                "water_footprint_country_average": {
                                    "green": row[i].value,
                                    "blue": excel_rows[idx + 1][i].value,
                                    "grey": excel_rows[idx + 2][i].value
                                }
                            }
                        )
        product["countries"] and products.append(product)
    return products

collection.insert_many(create_product_list(rows))
