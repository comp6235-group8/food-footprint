from openpyxl import load_workbook
from pymongo import MongoClient

client = MongoClient()
db = client["water_footprint"]
collection = db["crop_products"]
collection.delete_many({})

# Load excel file and select sheet
wb = load_workbook('./data/Report47-Appendix-II.xlsx')
ws = wb["App-II-WF_perTon"]

# Select all countries and column labels
countries = ws[4]
column_labels = ws[5]

# Rows we wish to loop over
rows = ws[7:994+2]


def create_product_list(excel_rows):
    """
    Creates a list of dictionaries representing a crop and its water footprint.
    Water footprint is represented at a global average + country averages.
    Includes both green, blue, and grey water footprint types.
    :param excel_rows: Excel rows to iterate through
    :return: A list of crop products
    """
    products = []
    for idx, row in enumerate(excel_rows):
        product = {"countries": []}
        if row[1].value:
            for i in range(len(column_labels)):
                if column_labels[i].value == "Product description (HS)":
                    product["product"] = row[i].value
                elif column_labels[i].value == "Product fraction (pf)":
                    product["product_fraction"] = row[i].value
                elif column_labels[i].value == "Value fraction (vf)":
                    product["value_fraction"] = row[i].value
                elif column_labels[i].value == "Global average":
                    product["water_footprint_global_average"] = {
                        "green": row[i].value,
                        "blue": excel_rows[idx + 1][i].value,
                        "grey": excel_rows[idx + 2][i].value
                    }
                elif column_labels[i].value == "CNTRY-average":
                    product["countries"].append(
                        {
                            "country": countries[i].value,
                            "water_footprint_country_average": {
                                "green": row[i].value,
                                "blue": excel_rows[idx + 1][i].value,
                                "grey": excel_rows[idx + 2][i].value
                            }
                        }
                    )
        product["countries"] and products.append(product)
    return products

collection.insert_many(create_product_list(rows))
